# -*- coding: utf-8 -*-
"""
CSV / Excel 导出模块
支持 UTF-8-BOM 编码（Excel兼容中文）
"""

import csv
import logging
from typing import List, Dict

from config import EXPORT_FIELDNAMES, EXPORT_HEADERS, CSV_PATH, EXCEL_PATH

logger = logging.getLogger(__name__)


def export_to_csv(data_list: List[Dict], filepath: str = None) -> str:
    """
    导出数据到 CSV 文件
    使用 UTF-8-BOM 编码，确保 Excel 正确显示中文

    Args:
        data_list: 数据列表，每个元素是一个字典
        filepath: 输出文件路径，默认使用配置中的路径

    Returns:
        输出文件路径
    """
    filepath = filepath or CSV_PATH

    if not data_list:
        logger.warning("[CSV导出] 数据为空，跳过导出")
        return filepath

    try:
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            # 使用中文表头
            writer = csv.DictWriter(f, fieldnames=EXPORT_FIELDNAMES, extrasaction='ignore')
            # 写入表头（中文）
            header_dict = dict(zip(EXPORT_FIELDNAMES, EXPORT_HEADERS))
            writer.writerow(header_dict)
            # 写入数据
            for row in data_list:
                writer.writerow(row)

        logger.info(f"[CSV导出] 成功导出 {len(data_list)} 条数据到 {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"[CSV导出] 导出失败: {e}")
        return filepath


def export_to_excel(data_list: List[Dict], filepath: str = None) -> str:
    """
    导出数据到 Excel 文件
    使用 openpyxl，设置列宽自适应和表头样式

    Args:
        data_list: 数据列表
        filepath: 输出文件路径

    Returns:
        输出文件路径
    """
    filepath = filepath or EXCEL_PATH

    if not data_list:
        logger.warning("[Excel导出] 数据为空，跳过导出")
        return filepath

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "车型数据"

        # 写入表头
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, header in enumerate(EXPORT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写入数据
        data_alignment = Alignment(vertical="center", wrap_text=True)
        for row_idx, row_data in enumerate(data_list, 2):
            for col_idx, field_name in enumerate(EXPORT_FIELDNAMES, 1):
                value = row_data.get(field_name, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border

        # 设置列宽自适应
        for col_idx, field_name in enumerate(EXPORT_FIELDNAMES, 1):
            max_length = len(EXPORT_HEADERS[col_idx - 1])
            for row_data in data_list[:100]:  # 只取前100行计算宽度
                value = str(row_data.get(field_name, ''))
                # 中文字符占2个宽度
                char_len = sum(2 if ord(c) > 127 else 1 for c in value)
                if char_len > max_length:
                    max_length = char_len
            # 限制最大宽度
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_length + 2, 40)

        # 冻结首行
        ws.freeze_panes = 'A2'

        # 自动筛选
        ws.auto_filter.ref = f"A1:{ws.cell(row=1, column=len(EXPORT_FIELDNAMES)).column_letter}{len(data_list) + 1}"

        wb.save(filepath)
        logger.info(f"[Excel导出] 成功导出 {len(data_list)} 条数据到 {filepath}")
        return filepath

    except ImportError:
        logger.warning("[Excel导出] openpyxl 未安装，跳过 Excel 导出。请运行: pip install openpyxl")
        return filepath
    except Exception as e:
        logger.error(f"[Excel导出] 导出失败: {e}")
        return filepath
